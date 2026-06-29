"""
file_handlers.py
================
High-performance file processing module for the Professional Spreadsheet Merger.
Handles streaming reads, delimiter detection, header extraction, and row iteration
for all supported formats while minimizing memory usage.

Supported formats:
  .csv .txt .prn .xls .xlsx .ods .xml .dif .slk
"""

import csv
import io
import os
import time
import xml.etree.ElementTree as ET
from typing import Dict, Iterator, List, Optional, Tuple

import pandas as pd

# ---------------------------------------------------------------------------
# Supported extensions (always compared lowercase)
# ---------------------------------------------------------------------------
SUPPORTED_EXTENSIONS: set = {
    ".csv",
    ".txt",
    ".xls",
    ".xlsx",
    ".xml",
    ".ods",
    ".dif",
    ".slk",
    ".prn",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_file_extension(filename: str) -> str:
    """Return the lowercase file extension including the leading dot."""
    return os.path.splitext(filename)[1].lower()


def is_supported_file(filename: str) -> bool:
    """Check whether the file extension is supported (case-insensitive)."""
    return get_file_extension(filename) in SUPPORTED_EXTENSIONS


def format_file_size(size_bytes: int) -> str:
    """Human-readable file size."""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.1f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"


def format_duration(seconds: float) -> str:
    """Format seconds into mm:ss or hh:mm:ss."""
    if seconds < 60:
        return f"{seconds:.1f}s"
    elif seconds < 3600:
        m, s = divmod(int(seconds), 60)
        return f"{m}m {s}s"
    else:
        h, rem = divmod(int(seconds), 3600)
        m, s = divmod(rem, 60)
        return f"{h}h {m}m {s}s"


# ---------------------------------------------------------------------------
# Delimiter detection
# ---------------------------------------------------------------------------

def detect_delimiter(file_path: str, sample_bytes: int = 16384) -> str:
    """
    Auto-detect the most likely delimiter for a CSV/TXT file by inspecting
    an initial byte sample.
    """
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as fh:
            sample = fh.read(sample_bytes)
    except Exception:
        with open(file_path, "r", encoding="latin-1", errors="replace") as fh:
            sample = fh.read(sample_bytes)

    candidates = [",", "\t", ";", "|", ":"]
    counts = {d: sample.count(d) for d in candidates}
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","


# ---------------------------------------------------------------------------
# Header extraction (fast, minimal I/O)
# ---------------------------------------------------------------------------

def extract_header(file_path: str) -> List[str]:
    """
    Extract the header row from any supported file format.
    Returns a list of column name strings.
    """
    ext = get_file_extension(file_path)

    # ------------------------------------------------------------------
    # CSV / TXT / PRN  ->  pandas read_csv with nrows=0
    # ------------------------------------------------------------------
    if ext in (".csv", ".txt", ".prn"):
        sep = detect_delimiter(file_path)
        if ext == ".prn":
            sep = r"\s+"
        df = pd.read_csv(
            file_path,
            nrows=0,
            sep=sep,
            engine="python" if ext == ".prn" else "c",
            encoding="utf-8",
            encoding_errors="replace",
        )
        return [str(c) for c in df.columns]

    # ------------------------------------------------------------------
    # XLSX  ->  openpyxl read-only, first row of first sheet
    # ------------------------------------------------------------------
    if ext == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        first = next(sheet.iter_rows(values_only=True), ())
        wb.close()
        return [str(cell) if cell is not None else "" for cell in first]

    # ------------------------------------------------------------------
    # XLS  ->  xlrd first row
    # ------------------------------------------------------------------
    if ext == ".xls":
        import xlrd

        book = xlrd.open_workbook(file_path, on_demand=True)
        sheet = book.sheet_by_index(0)
        header = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
        book.release_resources()
        return header

    # ------------------------------------------------------------------
    # ODS  ->  pandas read_excel(nrows=0) via odfpy
    # ------------------------------------------------------------------
    if ext == ".ods":
        df = pd.read_excel(file_path, engine="odf", nrows=0, dtype=str)
        return [str(c) for c in df.columns]

    # ------------------------------------------------------------------
    # XML  ->  pandas read_xml(nrows=0) or fallback to ElementTree
    # ------------------------------------------------------------------
    if ext == ".xml":
        try:
            df = pd.read_xml(file_path, nrows=0, dtype=str)
            return [str(c) for c in df.columns]
        except Exception:
            tree = ET.parse(file_path)
            root = tree.getroot()
            for elem in root.iter():
                children = list(elem)
                if children:
                    return [child.tag for child in children]
            return []

    # ------------------------------------------------------------------
    # DIF  ->  lightweight custom parser
    # ------------------------------------------------------------------
    if ext == ".dif":
        df = _parse_dif(file_path)
        return [str(c) for c in df.columns] if not df.empty else []

    # ------------------------------------------------------------------
    # SLK  ->  sylk_parser to CSV then read header
    # ------------------------------------------------------------------
    if ext == ".slk":
        from sylk_parser import SylkParser

        parser = SylkParser(file_path)
        buf = io.StringIO()
        parser.to_csv(buf)
        buf.seek(0)
        df = pd.read_csv(buf, nrows=0, dtype=str)
        return [str(c) for c in df.columns]

    return []


# ---------------------------------------------------------------------------
# Row count estimation (for UI statistics)
# ---------------------------------------------------------------------------

def estimate_row_count(file_path: str) -> int:
    """
    Return a rough estimate of data rows (excluding header).
    For text files this uses line-count heuristics; for binary formats
    the native APIs report exact sheet dimensions where possible.
    """
    ext = get_file_extension(file_path)

    if ext in (".csv", ".txt", ".prn"):
        try:
            with open(file_path, "rb") as fh:
                sample = fh.read(20480)
                if not sample:
                    return 0
                newline_count = sample.count(b"\n")
                avg_line = len(sample) / max(newline_count, 1)
                total_size = os.path.getsize(file_path)
                est = int(total_size / avg_line)
                return max(0, est - 1)
        except Exception:
            return 0

    if ext == ".xlsx":
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            total = sum(max(0, sheet.max_row - 1) for sheet in wb.worksheets)
            wb.close()
            return total
        except Exception:
            return 0

    if ext == ".xls":
        try:
            import xlrd

            book = xlrd.open_workbook(file_path, on_demand=True)
            total = sum(max(0, sheet.nrows - 1) for sheet in book.sheets())
            book.release_resources()
            return total
        except Exception:
            return 0

    if ext == ".ods":
        try:
            from odf import opendocument
            from odf.table import Table, TableRow

            doc = opendocument.load(file_path)
            tables = doc.spreadsheet.getElementsByType(Table)
            total = 0
            for table in tables:
                rows = list(table.getElementsByType(TableRow))
                total += max(0, len(rows) - 1)
            return total
        except Exception:
            return 0

    if ext == ".xml":
        try:
            count = 0
            context = ET.iterparse(file_path, events=("end",))
            for _, elem in context:
                tag = elem.tag.lower()
                if tag in ("row", "record", "item", "entry"):
                    count += 1
                elem.clear()
            return max(0, count - 1)
        except Exception:
            return 0

    if ext in (".dif", ".slk"):
        try:
            if ext == ".dif":
                df = _parse_dif(file_path)
            else:
                from sylk_parser import SylkParser

                parser = SylkParser(file_path)
                buf = io.StringIO()
                parser.to_csv(buf)
                buf.seek(0)
                df = pd.read_csv(buf, dtype=str)
            return max(0, len(df) - 1)
        except Exception:
            return 0

    return 0


# ---------------------------------------------------------------------------
# Streaming row generators
# ---------------------------------------------------------------------------

def stream_rows(
    file_path: str, master_columns: List[str], chunk_size: int = 50000
) -> Iterator[Dict[str, str]]:
    """
    Yield each row from *file_path* as a dictionary aligned to
    *master_columns*.  Missing columns receive an empty string.

    This generator is the heart of the memory-efficient merge:
    only one chunk (or one row) resides in RAM at any moment.
    """
    ext = get_file_extension(file_path)

    # ------------------------------------------------------------------
    # CSV / TXT  ->  chunked pandas read_csv
    # ------------------------------------------------------------------
    if ext in (".csv", ".txt"):
        sep = detect_delimiter(file_path)
        reader = pd.read_csv(
            file_path,
            chunksize=chunk_size,
            sep=sep,
            engine="c",
            dtype=str,
            keep_default_na=False,
            encoding="utf-8",
            encoding_errors="replace",
        )
        for chunk in reader:
            for _, row in chunk.iterrows():
                yield {col: str(row.get(col, "")) for col in master_columns}
        return

    # ------------------------------------------------------------------
    # PRN  ->  whitespace-delimited chunked read
    # ------------------------------------------------------------------
    if ext == ".prn":
        reader = pd.read_csv(
            file_path,
            chunksize=chunk_size,
            delim_whitespace=True,
            engine="python",
            dtype=str,
            keep_default_na=False,
            encoding="utf-8",
            encoding_errors="replace",
        )
        for chunk in reader:
            for _, row in chunk.iterrows():
                yield {col: str(row.get(col, "")) for col in master_columns}
        return

    # ------------------------------------------------------------------
    # XLSX  ->  openpyxl read-only row iterator
    # ------------------------------------------------------------------
    if ext == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                file_header = next(rows)
            except StopIteration:
                continue
            file_header = [str(h) if h is not None else "" for h in file_header]
            col_idx = {h: i for i, h in enumerate(file_header)}

            for row in rows:
                yield {
                    col: str(row[col_idx[col]])
                    if col in col_idx and col_idx[col] < len(row) and row[col_idx[col]] is not None
                    else ""
                    for col in master_columns
                }
        wb.close()
        return

    # ------------------------------------------------------------------
    # XLS  ->  xlrd sheet-by-sheet
    # ------------------------------------------------------------------
    if ext == ".xls":
        import xlrd

        book = xlrd.open_workbook(file_path, on_demand=True)
        for sheet_idx in range(book.nsheets):
            sheet = book.sheet_by_index(sheet_idx)
            if sheet.nrows == 0:
                continue
            file_header = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            col_idx = {h: c for c, h in enumerate(file_header)}

            for r in range(1, sheet.nrows):
                yield {
                    col: str(sheet.cell_value(r, col_idx[col]))
                    if col in col_idx
                    else ""
                    for col in master_columns
                }
        book.release_resources()
        return

    # ------------------------------------------------------------------
    # ODS  ->  pandas (ods are typically small; full load is acceptable)
    # ------------------------------------------------------------------
    if ext == ".ods":
        df = pd.read_excel(
            file_path, engine="odf", dtype=str, keep_default_na=False
        )
        for _, row in df.iterrows():
            yield {col: str(row.get(col, "")) for col in master_columns}
        return

    # ------------------------------------------------------------------
    # XML  ->  pandas read_xml
    # ------------------------------------------------------------------
    if ext == ".xml":
        df = pd.read_xml(file_path, dtype=str, keep_default_na=False)
        for _, row in df.iterrows():
            yield {col: str(row.get(col, "")) for col in master_columns}
        return

    # ------------------------------------------------------------------
    # DIF  ->  custom parser
    # ------------------------------------------------------------------
    if ext == ".dif":
        df = _parse_dif(file_path)
        for _, row in df.iterrows():
            yield {col: str(row.get(col, "")) for col in master_columns}
        return

    # ------------------------------------------------------------------
    # SLK  ->  sylk_parser
    # ------------------------------------------------------------------
    if ext == ".slk":
        from sylk_parser import SylkParser

        parser = SylkParser(file_path)
        buf = io.StringIO()
        parser.to_csv(buf)
        buf.seek(0)
        df = pd.read_csv(buf, dtype=str, keep_default_na=False)
        for _, row in df.iterrows():
            yield {col: str(row.get(col, "")) for col in master_columns}
        return


# ---------------------------------------------------------------------------
# DIF parser
# ---------------------------------------------------------------------------

def _parse_dif(file_path: str) -> pd.DataFrame:
    """
    Lightweight Data Interchange Format (DIF) parser.

    DIF structure (simplified):
      TABLE
      VECTORS
      TUPLES
      DATA
      0,0 ""
      -1,0 BOT
      <value rows>
      -1,0 EOD

    String values:  1,0"text"
    Numeric values: 0,number
    Row separator:  -1,0  (followed by BOT or EOD)
    """
    with open(file_path, "r", encoding="utf-8", errors="replace") as fh:
        lines = [ln.strip() for ln in fh.readlines()]

    # Locate DATA section
    data_start = None
    for i, line in enumerate(lines):
        if line.upper() == "DATA":
            data_start = i
            break

    if data_start is None:
        return pd.DataFrame()

    rows: List[List[str]] = []
    current: List[str] = []
    idx = data_start + 1

    while idx < len(lines):
        line = lines[idx]

        if line.upper() == "EOD":
            if current:
                rows.append(current)
            break

        if line == "-1,0":
            if current:
                rows.append(current)
                current = []
            idx += 1
            continue

        # String token
        if line.startswith("1,0"):
            val = line[3:].strip()
            if val.startswith('"') and val.endswith('"'):
                val = val[1:-1]
            current.append(val)
        # Numeric / special token
        elif line.startswith("0,"):
            parts = line.split(",", 1)
            if len(parts) > 1:
                val = parts[1].strip()
                if val == "V":
                    current.append("")
                else:
                    try:
                        current.append(str(float(val)))
                    except ValueError:
                        current.append(val)
            else:
                current.append("")
        else:
            current.append(line)

        idx += 1

    if not rows:
        return pd.DataFrame()

    header = rows[0]
    data = rows[1:]
    n_cols = len(header)

    # Normalise row lengths
    normalised = []
    for row in data:
        if len(row) < n_cols:
            row = row + [""] * (n_cols - len(row))
        elif len(row) > n_cols:
            row = row[:n_cols]
        normalised.append(row)

    return pd.DataFrame(normalised, columns=header)


# ---------------------------------------------------------------------------
# Master-column builder
# ---------------------------------------------------------------------------

def build_master_columns(file_infos: List[dict]) -> List[str]:
    """
    Scan every file's header and build a single master column list.
    Order is preserved from the first file; new columns discovered in
    later files are appended.
    """
    master: List[str] = []
    seen: set = set()

    for info in file_infos:
        try:
            header = extract_header(info["path"])
            for col in header:
                if col not in seen:
                    seen.add(col)
                    master.append(col)
        except Exception:
            continue

    return master


# ---------------------------------------------------------------------------
# Merge orchestrator
# ---------------------------------------------------------------------------

def perform_merge(
    file_infos: List[dict],
    master_columns: List[str],
    output_path: str,
    chunk_size: int = 50000,
    progress_callback=None,
) -> dict:
    """
    Stream-merge all files into a single CSV at *output_path*.

    Parameters
    ----------
    file_infos : list of dict
        Each dict must contain at least 'path' and 'name' keys.
    master_columns : list of str
        The canonical column order for the output CSV.
    output_path : str
        Destination file path.
    chunk_size : int
        Chunk size for pandas-based readers.
    progress_callback : callable or None
        If provided, called repeatedly with a dict of progress stats.

    Returns
    -------
    dict
        merge_result keys:
        - merged_files   : list of successfully merged filenames
        - skipped_files  : list of (filename, error_message) tuples
        - total_rows     : total data rows written
        - elapsed_seconds: wall-clock merge time
        - output_path    : path to the generated CSV
    """
    total_files = len(file_infos)
    start_time = time.time()

    merged_files: List[str] = []
    skipped_files: List[Tuple[str, str]] = []
    total_rows = 0

    with open(output_path, "w", newline="", encoding="utf-8") as out_fh:
        writer = csv.DictWriter(
            out_fh,
            fieldnames=master_columns,
            extrasaction="ignore",
            restval="",
            lineterminator="\n",
        )
        writer.writeheader()

        for file_idx, info in enumerate(file_infos):
            file_path = info["path"]
            file_name = info["name"]

            if progress_callback:
                progress_callback(
                    {
                        "current_file": file_name,
                        "file_index": file_idx,
                        "total_files": total_files,
                        "rows_processed": total_rows,
                        "elapsed": time.time() - start_time,
                    }
                )

            try:
                file_rows = 0
                for row in stream_rows(file_path, master_columns, chunk_size):
                    writer.writerow(row)
                    total_rows += 1
                    file_rows += 1

                    # Throttle UI updates to every 2 000 rows
                    if total_rows % 2000 == 0 and progress_callback:
                        progress_callback(
                            {
                                "current_file": file_name,
                                "file_index": file_idx,
                                "total_files": total_files,
                                "rows_processed": total_rows,
                                "elapsed": time.time() - start_time,
                            }
                        )

                merged_files.append(file_name)

            except Exception as exc:
                skipped_files.append((file_name, str(exc)))
                continue

    elapsed = time.time() - start_time

    return {
        "merged_files": merged_files,
        "skipped_files": skipped_files,
        "total_rows": total_rows,
        "elapsed_seconds": elapsed,
        "output_path": output_path,
    }
