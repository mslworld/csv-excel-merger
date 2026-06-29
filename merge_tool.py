"""
Professional CSV/Excel/Spreadsheet Merger
==========================================
A high-performance Streamlit web application for merging unlimited spreadsheet
files into a single CSV using streaming I/O, chunked processing, and minimal RAM usage.

Supported formats: .csv, .txt, .xls, .xlsx, .xml, .ods, .dif, .slk, .prn
All extensions are case-insensitive.

Author: Professional Data Tools
"""

import csv
import io
import os
import tempfile
import time
import uuid
import xml.etree.ElementTree as ET
from typing import Dict, Iterator, List

import pandas as pd
import streamlit as st

# =============================================================================
# Configuration
# =============================================================================
FILES_PER_PAGE = 20
DEFAULT_CHUNK_SIZE = 50000
SUPPORTED_EXTENSIONS = {".csv", ".txt", ".xls", ".xlsx", ".xml", ".ods", ".dif", ".slk", ".prn"}

# =============================================================================
# Page Setup
# =============================================================================
st.set_page_config(
    page_title="Professional Spreadsheet Merger",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Custom CSS
st.markdown(
    """
    <style>
        .block-container { padding-top: 1.5rem; padding-bottom: 1.5rem; }
        .file-row { display: flex; align-items: center; padding: 0.5rem 0.75rem; border-bottom: 1px solid #e5e7eb; font-size: 0.9rem; }
        .file-row:hover { background-color: #f9fafb; }
        .metric-card { background: #ffffff; border: 1px solid #e5e7eb; border-radius: 0.5rem; padding: 1rem; text-align: center; }
        .metric-value { font-size: 1.5rem; font-weight: 700; color: #111827; }
        .metric-label { font-size: 0.75rem; color: #6b7280; text-transform: uppercase; letter-spacing: 0.05em; }
        .stButton>button { border-radius: 0.375rem; font-weight: 500; }
    </style>
    """,
    unsafe_allow_html=True,
)

# =============================================================================
# Helper Functions
# =============================================================================

def get_extension(filename: str) -> str:
    return os.path.splitext(filename)[1].lower()

def is_supported(filename: str) -> bool:
    return get_extension(filename) in SUPPORTED_EXTENSIONS

def format_size(size_bytes: int) -> str:
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.1f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"

def format_duration(seconds: float) -> str:
    if seconds < 60:
        return f"{seconds:.1f}s"
    elif seconds < 3600:
        m, s = divmod(int(seconds), 60)
        return f"{m}m {s}s"
    else:
        h, rem = divmod(int(seconds), 3600)
        m, s = divmod(rem, 60)
        return f"{h}h {m}m {s}s"

def detect_delimiter(file_path: str, sample_bytes: int = 16384) -> str:
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as fh:
            sample = fh.read(sample_bytes)
    except Exception:
        with open(file_path, "r", encoding="latin-1", errors="replace") as fh:
            sample = fh.read(sample_bytes)
    candidates = [",", "\t", ";", "|", ":"]
    counts = {d: sample.count(d) for d in candidates}
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","

def extract_header(file_path: str) -> List[str]:
    ext = get_extension(file_path)
    try:
        if ext in (".csv", ".txt", ".prn"):
            sep = detect_delimiter(file_path) if ext in (".csv", ".txt") else r"\s+"
            df = pd.read_csv(file_path, nrows=0, sep=sep, engine="python" if ext == ".prn" else "c", encoding="utf-8", encoding_errors="replace")
            return [str(c) for c in df.columns]
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            first = next(sheet.iter_rows(values_only=True), ())
            wb.close()
            return [str(cell) if cell is not None else "" for cell in first]
        if ext == ".xls":
            import xlrd
            book = xlrd.open_workbook(file_path, on_demand=True)
            sheet = book.sheet_by_index(0)
            header = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            book.release_resources()
            return header
        if ext == ".ods":
            df = pd.read_excel(file_path, engine="odf", nrows=0, dtype=str)
            return [str(c) for c in df.columns]
        if ext == ".xml":
            try:
                df = pd.read_xml(file_path, nrows=0, dtype=str)
                return [str(c) for c in df.columns]
            except:
                return []
        if ext == ".slk":
            from sylk_parser import SylkParser
            parser = SylkParser(file_path)
            buf = io.StringIO()
            parser.to_csv(buf)
            buf.seek(0)
            df = pd.read_csv(buf, nrows=0, dtype=str)
            return [str(c) for c in df.columns]
    except Exception:
        pass
    return []

def estimate_rows(file_path: str) -> int:
    ext = get_extension(file_path)
    try:
        if ext in (".csv", ".txt", ".prn"):
            with open(file_path, "rb") as fh:
                sample = fh.read(20480)
                if not sample:
                    return 0
                nl = sample.count(b"\n")
                avg_line = len(sample) / max(nl, 1)
                return max(0, int(os.path.getsize(file_path) / avg_line) - 1)
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            total = sum(max(0, s.max_row - 1) for s in wb.worksheets)
            wb.close()
            return total
        if ext == ".xls":
            import xlrd
            book = xlrd.open_workbook(file_path, on_demand=True)
            total = sum(max(0, s.nrows - 1) for s in book.sheets())
            book.release_resources()
            return total
        if ext == ".ods":
            from odf import opendocument
            from odf.table import Table, TableRow
            doc = opendocument.load(file_path)
            tables = doc.spreadsheet.getElementsByType(Table)
            total = sum(max(0, len(list(t.getElementsByType(TableRow))) - 1) for t in tables)
            return total
        if ext == ".xml":
            count = 0
            for _, elem in ET.iterparse(file_path, events=("end",)):
                if elem.tag.lower() in ("row", "record", "item", "entry"):
                    count += 1
                elem.clear()
            return max(0, count - 1)
        if ext == ".dif":
            df = _parse_dif_py(file_path)
            return max(0, len(df) - 1)
        if ext == ".slk":
            from sylk_parser import SylkParser
            parser = SylkParser(file_path)
            buf = io.StringIO()
            parser.to_csv(buf)
            buf.seek(0)
            df = pd.read_csv(buf, dtype=str)
            return max(0, len(df) - 1)
    except:
        pass
    return 0

def stream_rows_positional(file_path: str, master_cols: List[str], chunk_size: int = DEFAULT_CHUNK_SIZE) -> Iterator[List[str]]:
    """Yield rows as lists of values aligned POSITIONALLY to master columns."""
    ext = get_extension(file_path)
    try:
        if ext in (".csv", ".txt"):
            sep = detect_delimiter(file_path)
            reader = pd.read_csv(file_path, chunksize=chunk_size, sep=sep, engine="c", dtype=str, keep_default_na=False, encoding="utf-8", encoding_errors="replace")
            for chunk in reader:
                for vals in chunk.values:
                    yield [str(v) if i < len(vals) else "" for i in range(len(master_cols))]
            return
        if ext == ".prn":
            reader = pd.read_csv(file_path, chunksize=chunk_size, delim_whitespace=True, engine="python", dtype=str, keep_default_na=False, encoding="utf-8", encoding_errors="replace")
            for chunk in reader:
                for vals in chunk.values:
                    yield [str(v) if i < len(vals) else "" for i in range(len(master_cols))]
            return
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet in wb.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                for row in rows[1:]:  # Skip header
                    vals = [str(c) if c is not None else "" for c in row]
                    yield [vals[i] if i < len(vals) else "" for i in range(len(master_cols))]
            wb.close()
            return
        if ext == ".xls":
            import xlrd
            book = xlrd.open_workbook(file_path, on_demand=True)
            for s_idx in range(book.nsheets):
                sheet = book.sheet_by_index(s_idx)
                for r in range(1, sheet.nrows):
                    vals = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                    yield [vals[i] if i < len(vals) else "" for i in range(len(master_cols))]
            book.release_resources()
            return
        if ext == ".ods":
            df = pd.read_excel(file_path, engine="odf", dtype=str, keep_default_na=False)
            for vals in df.values:
                yield [str(v) if i < len(vals) else "" for i in range(len(master_cols))]
            return
        if ext == ".xml":
            df = pd.read_xml(file_path, dtype=str, keep_default_na=False)
            for vals in df.values:
                yield [str(v) if i < len(vals) else "" for i in range(len(master_cols))]
            return
        if ext == ".slk":
            from sylk_parser import SylkParser
            parser = SylkParser(file_path)
            buf = io.StringIO()
            parser.to_csv(buf)
            buf.seek(0)
            df = pd.read_csv(buf, dtype=str, keep_default_na=False)
            for vals in df.values:
                yield [str(v) if i < len(vals) else "" for i in range(len(master_cols))]
            return
        if ext == ".dif":
            df = _parse_dif_py(file_path)
            for vals in df.values:
                yield [str(v) if i < len(vals) else "" for i in range(len(master_cols))]
            return
    except Exception as e:
        raise e

def _parse_dif_py(file_path: str) -> pd.DataFrame:
    with open(file_path, "r", encoding="utf-8", errors="replace") as fh:
        lines = [ln.strip() for ln in fh.readlines()]
    data_start = next((i for i, ln in enumerate(lines) if ln.upper() == "DATA"), None)
    if data_start is None:
        return pd.DataFrame()
    rows, current = [], []
    for line in lines[data_start + 1:]:
        if line.upper() == "EOD":
            if current: rows.append(current)
            break
        if line == "-1,0":
            if current: rows.append(current)
            current = []
            continue
        if line.startswith("1,0"):
            val = line[3:].strip()
            if val.startswith('"') and val.endswith('"'):
                val = val[1:-1]
            current.append(val)
        elif line.startswith("0,"):
            parts = line.split(",", 1)
            val = parts[1].strip() if len(parts) > 1 else ""
            current.append("" if val == "V" else val)
        else:
            current.append(line)
    if not rows:
        return pd.DataFrame()
    header, data = rows[0], rows[1:]
    n = len(header)
    normalised = [[row[i] if i < len(row) else "" for i in range(n)] for row in data]
    return pd.DataFrame(normalised, columns=header)

# =============================================================================
# Session State Init
# =============================================================================
if "files" not in st.session_state:
    st.session_state.files = []
if "merge_result" not in st.session_state:
    st.session_state.merge_result = None
if "temp_dir" not in st.session_state:
    st.session_state.temp_dir = tempfile.mkdtemp(prefix="merger_")

# =============================================================================
# Sidebar
# =============================================================================
with st.sidebar:
    st.image("https://img.icons8.com/color/96/spreadsheet.png", width=48)
    st.title("Settings")
    st.markdown("---")
    chunk_size = st.number_input("Chunk size (rows)", 1000, 500000, DEFAULT_CHUNK_SIZE, 5000,
                                  help="Rows per chunk for CSV/TXT. Larger = faster but more RAM.")
    st.markdown("---")
    st.subheader("Supported formats")
    st.markdown("- **CSV** – Comma-separated values\n- **TXT** – Plain text tables\n- **PRN** – Fixed-width\n- **XLS/XLSX** – Excel\n- **ODS** – OpenDocument\n- **XML** – Spreadsheet XML\n- **DIF** – Data Interchange\n- **SLK** – SYLK")
    st.markdown("---")
    st.caption("Built for speed. Streaming I/O keeps RAM usage flat.")

# =============================================================================
# Main UI
# =============================================================================
st.title("📊 Professional Spreadsheet Merger")
st.markdown("Merge unlimited files into a single CSV using **streaming technology**.")

uploaded = st.file_uploader("Drag & drop files or click to browse", type=list(SUPPORTED_EXTENSIONS), accept_multiple_files=True)
if uploaded:
    temp_dir = st.session_state.temp_dir
    for uf in uploaded:
        if not is_supported(uf.name):
            continue
        if any(f["name"] == uf.name for f in st.session_state.files):
            continue
        file_id = str(uuid.uuid4())[:8]
        dest = os.path.join(temp_dir, f"{file_id}_{uf.name}")
        with open(dest, "wb") as fh:
            fh.write(uf.getbuffer())
        st.session_state.files.append({
            "id": file_id, "name": uf.name, "path": dest,
            "size": os.path.getsize(dest), "ext": get_extension(uf.name),
            "rows": estimate_rows(dest)
        })

if not st.session_state.files:
    st.info("👆 Upload at least one supported file.")
    st.stop()

# Search & Remove All
search = st.text_input("🔍 Search files", placeholder="Type to filter...", label_visibility="collapsed")
files_filtered = [f for f in st.session_state.files if search.lower() in f["name"].lower()]

# Summary
summary = {"count": len(st.session_state.files), "size": sum(f["size"] for f in st.session_state.files),
           "rows": sum(f["rows"] for f in st.session_state.files)}
cols = st.columns(5)
cols[0].metric("Total Files", f"{summary['count']:,}")
cols[1].metric("Upload Size", format_size(summary["size"]))
cols[2].metric("Est. Rows", f"{summary['rows']:,}")
cols[3].metric("Unique Formats", len(set(f["ext"] for f in st.session_state.files)))
cols[4].metric("Top Format", max(set(f["ext"] for f in st.session_state.files), key=lambda x: sum(1 for f in st.session_state.files if f["ext"] == x)))

# Pagination
total_pages = max(1, (len(files_filtered) + FILES_PER_PAGE - 1) // FILES_PER_PAGE)
page = st.number_input("Page", 1, total_pages, 1, key="page_num")
start_idx = (page - 1) * FILES_PER_PAGE
end_idx = start_idx + FILES_PER_PAGE

st.markdown("### File List:")
header_cols = st.columns([4, 2, 1, 1])
header_cols[0].markdown("**Filename**")
header_cols[1].markdown("**Size**")
header_cols[2].markdown("**Ext**")
header_cols[3].markdown("**Action**")

for f in files_filtered[start_idx:end_idx]:
    row_cols = st.columns([4, 2, 1, 1])
    row_cols[0].text(f["name"])
    row_cols[1].text(format_size(f["size"]))
    row_cols[2].code(f["ext"])
    if row_cols[3].button("🗑️", key=f"del_{f['id']}"):
        os.remove(f["path"])
        st.session_state.files.remove(f)
        st.rerun()

# Remove All button
if st.button("🗑️ Remove All", type="secondary"):
    for f in st.session_state.files:
        try:
            os.remove(f["path"])
        except:
            pass
    st.session_state.files = []
    st.session_state.merge_result = None
    st.rerun()

st.markdown("---")

# Merge Button
if st.button("🚀 Merge Files", type="primary", use_container_width=True):
    st.session_state.merge_result = None
    
    # Phase 1: Build master columns
    with st.status("🔍 Scanning headers...", expanded=True) as status:
        master_cols = []
        seen = set()
        for f in st.session_state.files:
            try:
                h = extract_header(f["path"])
                for col in h:
                    if col not in seen:
                        seen.add(col)
                        master_cols.append(col)
            except:
                pass
        status.update(label=f"✅ Found {len(master_cols)} unique columns", state="complete")
    
    if not master_cols:
        st.error("No columns found in any file.")
        st.stop()
    
    # Phase 2: Streaming merge
    output_path = os.path.join(st.session_state.temp_dir, "merged_output.csv")
    progress_bar = st.progress(0.0)
    status_text = st.empty()
    stats_text = st.empty()
    
    start_time = time.time()
    merged_files, skipped_files, total_rows = [], [], 0
    
    with open(output_path, "w", newline="", encoding="utf-8") as out_fh:
        writer = csv.writer(out_fh, lineterminator="\n")
        writer.writerow(master_cols)
        
        for idx, f in enumerate(st.session_state.files):
            if progress_bar:
                progress_bar.progress((idx + 0.5) / len(st.session_state.files))
                status_text.markdown(f"**Current:** `{f['name']}`  \n**Files:** {idx+1}/{len(st.session_state.files)}")
            
            try:
                for row in stream_rows_positional(f["path"], master_cols, chunk_size):
                    writer.writerow(row)
                    total_rows += 1
                    if total_rows % 5000 == 0:
                        elapsed = time.time() - start_time
                        eta = elapsed / (idx + 1) * len(st.session_state.files) - elapsed if idx > 0 else 0
                        stats_text.markdown(f"⏱️ {format_duration(elapsed)} elapsed | ⏳ ETA: {format_duration(max(0, eta))}")
                merged_files.append(f["name"])
            except Exception as e:
                skipped_files.append((f["name"], str(e)))
    
    elapsed = time.time() - start_time
    st.session_state.merge_result = {
        "merged_files": merged_files, "skipped_files": skipped_files,
        "total_rows": total_rows, "elapsed_seconds": elapsed, "output_path": output_path
    }
    st.session_state.merge_result = st.session_state.merge_result

# Results
if st.session_state.merge_result:
    r = st.session_state.merge_result
    st.balloons()
    st.success("✅ Merge completed!")
    
    stat_cols = st.columns(5)
    stat_cols[0].metric("Merged Files", len(r["merged_files"]))
    stat_cols[1].metric("Skipped", len(r["skipped_files"]))
    stat_cols[2].metric("Total Rows", f"{r['total_rows']:,}")
    stat_cols[3].metric("Time", format_duration(r["elapsed_seconds"]))
    stat_cols[4].metric("Output Size", format_size(os.path.getsize(r["output_path"])))
    
    with open(r["output_path"], "rb") as fh:
        st.download_button("⬇️ Download Merged CSV", fh, "merged_output.csv", "text/csv", use_container_width=True)
    
    if r["skipped_files"]:
        with st.expander(f"⚠️ Skipped Files ({len(r['skipped_files'])})"):
            for name, err in r["skipped_files"]:
                st.error(f"**{name}**: {err}")

st.markdown("---")
st.caption("Tip: Upload more files and merge again without reloading.")
